import os
import tempfile
import unittest
from unittest.mock import patch, MagicMock

import openpyxl as xl

from ofrom_outils.common import TRUNC
from ofrom_outils.stats.token_dict import TokenDict


def set_tmp(suffix=".xlsx"):
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_name = tmp.name
    return tmp_name


class TestTokenDict(unittest.TestCase):

    def setUp(self):
        self.td = TokenDict("not-a-path")
        self.wb = xl.Workbook()
        sh = self.wb.active
        sh.title = "b"
        sh.append(("mot", "lemme", "pos", "phones", "nb", "fichiers"))
        sh.append(("bateau", "bateau", "NOM:com", "", "2", "fi1,fi2"))

    @patch("ofrom_outils.stats.token_dict.iter_segs")
    @patch("ofrom_outils.stats.token_dict.fromPraat")
    @patch("ofrom_outils.stats.token_dict.iter_core")
    def test_iter_toks(self, mock_core, mock_praat, mock_segs):
        mock_core.return_value = [
            ("corp", "fi1", "ext", "fi1.ext", "path"),
            ("corp", "fi2", "ext", "fi2.ext", "path")
        ]
        mock_praat.fromPraat.return_value = "Trans"
        mock_segs.return_value = ["t1", "t2", "t3"]
        l_res = [tpl for tpl in self.td._iter_toks()]
        self.assertEqual(l_res, [
            ("corp", "fi1", "Trans", "t1"),
            ("corp", "fi1", "Trans", "t2"),
            ("corp", "fi1", "Trans", "t3"),
            ("corp", "fi2", "Trans", "t1"),
            ("corp", "fi2", "Trans", "t2"),
            ("corp", "fi2", "Trans", "t3"),
        ])
        assert mock_praat.fromPraat.call_count == 2

    def test_get_anno(self):
        l_atier = []
        for i in range(3):
            mtier = MagicMock()
            mtier.elem = []
            for j in range(3):
                maseg = MagicMock()
                maseg.index.return_value = j
                maseg.start = j
                maseg.end = j + 1.2
                maseg.content = f"{i}-{j}"
                mtier.elem.append(maseg)
            mtier.getTime.return_value = mtier.elem[0]
            l_atier.append(mtier)
        mseg = MagicMock()
        mseg.end = 1.5
        l_res = self.td._get_anno(mseg, l_atier)
        self.assertEqual(l_res, [
            '0-0,0-1', '1-0,1-1', '2-0,2-1'
        ])

    def test_get_row_dict(self):
        mock_sh = MagicMock()
        l_res = []
        for i in range(4):
            mock_cell = MagicMock()
            mock_cell.value = f"a{i}"
            l_res.append(("buffer_row", mock_cell, "col2", "col3"))
        mock_sh.iter_rows.return_value = l_res
        d_res = self.td.get_row_dict(mock_sh)
        self.assertEqual(d_res, {
            'a0': 2, 'a1': 3, 'a2': 4, 'a3': 5
        })

    def test_ofrom_dict_c(self):
        self.assertEqual(self.td.ofrom_dict_c("bateau"), "b")
        self.assertEqual(self.td.ofrom_dict_c("élan"), "e")
        self.assertEqual(self.td.ofrom_dict_c("Bateau"), "autre")
        self.assertEqual(self.td.ofrom_dict_c("#"), None)
        self.assertEqual(self.td.ofrom_dict_c(f"bate{TRUNC}"), "tronc")
        self.assertEqual(self.td.ofrom_dict_c("bate/"), "tronc")

    @patch("ofrom_outils.stats.token_dict.TokenDict._get_anno")
    @patch("ofrom_outils.stats.token_dict.TokenDict._iter_toks")
    def test_generate(self, mock_iter, mock_anno):
        td = TokenDict()
        l_iter: list[str | tuple[str, str, MagicMock, str]] = \
            ["bateau", "élan", "Bateau", "bateau", f"bat{TRUNC}"]
        for i in range(5):
            mtrans = MagicMock()
            mtrans.getName.return_value = "tier"
            mtok = MagicMock()
            mtok.struct.return_value = MagicMock()
            mstruct = mtok.struct.return_value
            mstruct.name = "spk"
            mtok.content = l_iter[i]
            l_iter[i] = ("corp", f"fi{1 + int(i / 2)}", mtrans, mtok)
        mock_iter.return_value = l_iter
        mock_anno.return_value = ["pos", "lem", ["phn"]]
        d_res = td.generate()
        d_res = {k: v for k, v in d_res.items() if v != {}}
        self.assertEqual(d_res, {
            'autre': {
                'Bateau': {
                    'pos': ['pos'], 'lemma': ['lem'], 'phones': ['phn'],
                    'nb': 1, 'files': ['fi2']
                },
            },
            'tronc': {
                'bat-': {
                    'pos': ['pos'], 'lemma': ['lem'], 'phones': ['phn'],
                    'nb': 1, 'files': ['fi3']
                }
            },
            'b': {
                'bateau': {
                    'pos': ['pos'], 'lemma': ['lem'], 'phones': ['phn'],
                    'nb': 2, 'files': ['fi1', 'fi2']
                }
            },
            'e': {
                'élan': {
                    'pos': ['pos'], 'lemma': ['lem'], 'phones': ['phn'],
                    'nb': 1, 'files': ['fi1']
                }
            }
        })

    @patch("ofrom_outils.stats.token_dict.xl")
    @patch("ofrom_outils.stats.token_dict.os.path.isfile")
    def test_load(self, mock_isfile, mock_xl):
        mock_isfile.return_value = True
        mock_xl.load_workbook.return_value = self.wb
        d_res = self.td.load()
        self.assertEqual(d_res, {
            'b': {
                'bateau': {
                    'lemme': ['bateau'], 'pos': ['NOM:com'], 'phones': [''],
                    'nb': 2, 'fichiers': ['fi1', 'fi2']
                }
            }
        })

    def test_save(self):
        self.td.d = {
            'b': {
                'bateau': {
                    'lemme': ['bateau'], 'pos': ['NOM:com'], 'phones': [''],
                    'nb': 2, 'fichiers': ['fi1', 'fi2']
                }
            }
        }
        self.td.f = set_tmp()
        try:
            self.td.save()
            wb = xl.load_workbook(self.td.f)
            sh = wb['b']
            word = sh.cell(row=2, column=1).value
            self.assertIn("autre", wb.sheetnames)
            self.assertEqual(sh.title, "b")
            self.assertEqual(word, 'bateau')
        finally:
            os.remove(self.td.f)

    def test_get(self):
        self.td.d = {
            'b': {
                'bateau': {
                    'lemme': ['bateau'], 'pos': ['NOM:com'], 'phones': [''],
                    'nb': 2, 'fichiers': ['fi1', 'fi2']
                }
            }
        }
        res = self.td.get("voilier")
        self.assertEqual(res, None)
        res = self.td.get("bateau")
        self.assertEqual(res, {
            'lemme': ['bateau'], 'pos': ['NOM:com'], 'phones': [''],
            'nb': 2, 'fichiers': ['fi1', 'fi2']
        })