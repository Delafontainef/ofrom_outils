import os
import re

import openpyxl as xl
from corflow import fromPraat
from corflow.Transcription import Transcription, Tier, Segment
from openpyxl.worksheet.worksheet import Worksheet

from ofrom_outils.common import (
    ROOT, TAGS, SYMS, iter_core, iter_segs
)
from ofrom_outils.common_types import Path, Iterator
from ofrom_outils.logs import log
from ofrom_outils.stats.stats_models import OFRDict, AbsTokenDict

"""Constantes globales
clé         type        description
-------------------------------------------------
L_SH        list<str>   nom des feuilles du dictionnaire d'OFROM+.
D_TR        dict        conversion de lettres pour 'L_SH'.
PH, ...     str         suffixes de tire pour le type d'annotation.
HEAD        list<str>   en-tête de feuillet pour fichier Excel.
"""
L_SH = ["autre", "tronc", "a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l",
        "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]
D_TR = {
    'à': 'a', 'â': 'a',
    'é': 'e', 'è': 'e', 'ê': 'e',
    'î': 'i', 'ì': 'i',
    'ô': 'o', 'œ': 'o',
    'ç': 'c'
}
PH, WD = TAGS['phn'], TAGS['wrd']
TS, PS, LE = TAGS['tok'], TAGS['pos'], TAGS['lem']
HEAD = ["mot", "lemme", "pos", "phones", "nb", "fichiers"]


class TokenDict(AbsTokenDict):

    def __init__(self):
        self.f = os.path.join(ROOT, "ofrom_outils", "stats",
                              "ofrom_dict.xlsx")
        self.d = {}

        # Méthodes privées #
        # ------------------#

    @staticmethod
    def _iter_toks() -> Iterator[tuple[str, str, Transcription, Segment]]:
        """Génère un dictionnaire de tokens du corpus OFROM+."""
        for corp, fi, ext, file, path in iter_core():
            log(fi)
            trans = fromPraat.fromPraat(path)
            for tok in iter_segs(trans, [TS, WD]):
                yield corp, fi, trans, tok

    @staticmethod
    def _get_anno(seg: Segment, l_atier: list[Tier]) -> list[str]:
        """Retourne les annotations sous forme de liste."""
        l_res = []
        for atier in l_atier:
            aseg = atier.getTime(seg.start) if atier else None
            if not aseg:
                l_res.append("")
                continue
            ac = aseg.content
            ac = ac.split("|", 1)[0] if "|" in ac else ac  # lemmas
            l_ac = [ac]
            while aseg.end < seg.end:  # for phones (or multipart annotation)
                aseg = atier.elem[aseg.index() + 1]
                ac = aseg.content
                ac = ac.split("|", 1)[0] if "|" in ac else ac
                l_ac.append(ac)
            l_res.append(",".join(l_ac) if len(l_ac) > 1 else l_ac[0])
        return l_res

    @staticmethod
    def _get_row_dict(sh: Worksheet) -> dict[str, int]:
        """Lignes par token."""
        d_sh = {}
        a = 2
        for row in sh.iter_rows(2):
            d_sh[row[1].value] = a
            a += 1
        return d_sh

    @staticmethod
    def ofrom_dict_c(word: str) -> None | str:
        """Retourne la feuille de dictionnaire avec le token 'word'."""
        if (not word) or word in SYMS:  # invalid word
            return None
        c = word[0]  # first letter
        c = D_TR[c] if c in D_TR else c  # accents, etc.
        c = 'autre' if c not in L_SH else c  # dict' sheet
        c = 'tronc' if word.endswith("-") else c  # truncation
        return c

        # Gérer le dictionnaire #
        # -----------------------#

    def generate(self) -> OFRDict:
        """Génère un dictionnaire de tokens du corpus OFROM+."""
        self.d = {c: {} for c in L_SH}
        ostruct, spk = None, ""
        ps_tier, le_tier, ph_tier = None, None, None
        for corp, fi, trans, tok in self._iter_toks():
            if ostruct != tok.struct:
                ostruct = tok.struct
                spk = tok.struct.name.split("[", 1)[0]
                ps_tier = trans.getName(spk + PS)  # PoS
                if not ps_tier:  # "temporary" fix
                    ps_tier = trans.getName(spk + "[pos]")
                le_tier = trans.getName(spk + LE)  # lemma
                ph_tier = trans.getName(spk + PH)  # phones
            if (not tok.content) or re.search(SYMS, tok.content):
                continue  # ignore symbols
            c = tok.content[0]
            c = "tronc" if tok.content.endswith("-") else c
            c = D_TR[c] if c in D_TR else c
            c = "autre" if not c in self.d else c  # get 'c' key
            l_anno = self._get_anno(tok, [ps_tier, le_tier, ph_tier])
            if tok.content not in self.d[c]:  # new word
                self.d[c][tok.content] = {
                    "pos": [l_anno[0]],
                    "lemma": [l_anno[1]],
                    "phones": l_anno[2],
                    "nb": 1,
                    "files": [fi]
                }
            else:  # append
                self.d[c][tok.content]['nb'] += 1
                if l_anno[0] not in self.d[c][tok.content]['pos']:
                    self.d[c][tok.content]["pos"].append(l_anno[0])
                if l_anno[1] not in self.d[c][tok.content]['lemma']:
                    self.d[c][tok.content]["lemma"].append(l_anno[1])
                if trans.name not in self.d[c][tok.content]['files']:
                    self.d[c][tok.content]['files'].append(fi)
        return self.d

    def load(self, path: Path = "") -> OFRDict:
        """Charge le dictionnaire depuis un fichier Excel."""
        path = self.f if not path else path
        if not os.path.isfile(path):
            return {}
        self.d, wb = {}, xl.load_workbook(path)
        for shn in wb.sheetnames:
            self.d[shn], sh = {}, wb[shn]
            d_c = {c.value: i for i, c in enumerate(sh[1])}
            for i, row in enumerate(sh.iter_rows(min_row=2)):
                i = i + 2
                word = sh.cell(row=i, column=d_c['mot']).value
                self.d[shn][word] = {
                    'lemme': sh.cell(row=i, column=d_c['lemme']).value,
                    'pos': sh.cell(row=i, column=d_c['pos']).value,
                    'phones': sh.cell(row=i, column=d_c['phones']).value,
                    'nb': sh.cell(row=i, column=d_c['nb']).value,
                    'fichiers': sh.cell(row=i, column=d_c['fichiers']).value
                }
        return self.d

    def save(self, path: Path = "") -> None:
        """Sauvegarde sous forme de fichier Excel."""
        path = self.f if not path else path
        wb = xl.Workbook()
        sh = wb.active
        sh.title = "autre"
        sh.append(HEAD)
        for c in L_SH:
            if c == "autre":
                sh = wb["autre"]
            elif c not in wb.sheetnames:
                sh = wb.create_sheet()
                sh.title = c
                sh.append(HEAD)
            else:
                sh = wb[c]
            # d_sh = self._get_row_dict(sh)
            d_c = self.d[c]
            l_c = list(d_c.keys())
            l_c.sort()
            for k in l_c:
                d_v = d_c[k]
                l_p, l_le, ph = d_v['pos'], d_v['lemma'], d_v['phones']
                nb, l_fi = d_v['nb'], d_v['files']
                pos = ",".join(l_p)
                le = ",".join(l_le)
                vf = ",".join(l_fi) if len(l_fi) <= 10 else \
                    str(len(l_fi))  # ahem...
                sh.append([k, le, pos, ph, nb, vf])
        wb.save(path)

        # Utiliser le dictionnaire #
        # --------------------------#

    def get(self, word: str) -> None | dict[str, str]:
        """Récupère les informations sur 'word' à partir du dictionnaire 'd'.
           Renvoie un booléen et l'entrée du dictionnaire correspondante.
        """
        if (not word) or word in SYMS:
            return None
        c = self.ofrom_dict_c(word)
        try:
            return self.d[c][word]
        except KeyError:
            return None
