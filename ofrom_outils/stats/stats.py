"""28.02.2026
Script principal pour les statistiques.
Récupère les statistiques et les charge/sauvegarde dans un fichier Excel.
"""
import datetime
import os

import openpyxl as xl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ofrom_outils.common import fix_lext
from ofrom_outils.common_types import Any, Path, Callable
from ofrom_outils.meta.meta import Meta
from ofrom_outils.meta.meta_validation import CELL_D, CELL_Y
from ofrom_outils.stats.stats_models import StFile, StList, AbsStats
from ofrom_outils.stats.stats_reading import (
    fill_stlist, all_file_stats, corp_stats
)

SCOR = "sous-corpus"
HEADER = [SCOR, "nb_loc", "nb_enr", "nb_mots", "duree"]


def open_excel(path: Path, shn: str) -> tuple[Workbook, Worksheet]:
    """Ouvre un fichier Excel et sélectionne le feuillet."""
    fi, ext = os.path.splitext(path)
    if os.path.isfile(path) and ext.lower() == ".xlsx":
        wb = xl.load_workbook(path)
        if shn in wb.sheetnames:
            sh = wb[shn]
        else:
            wb.create_sheet(shn)
            sh = wb[shn]
    else:
        wb = xl.Workbook()
        sh: Worksheet = wb.active
        sh.title = shn
    return wb, sh


def _write_table(sh: Worksheet, d_st: dict[str, StList], head: list[str]
                 ) -> tuple[int, int, int, float]:
    """Crée une nouvelle table."""
    head = HEADER if not head else head
    sh.append([])
    sh.append(head)
    genr, gwd, gdur = 0, 0, 0.0
    s_spk = set()
    for typ, st in sorted(d_st.items()):
        loc, enr, wd, dur = len(st.spk), len(st.fi), st.wd, st.dur
        sh.append([typ, loc, enr, wd, dur])
        genr += enr
        gwd += wd
        gdur += dur
        s_spk.update(st.spk)
    gloc = len(s_spk)
    sh.append(["Total", gloc, genr, gwd, gdur])
    return gloc, genr, gwd, gdur


class Stats(AbsStats):

    def __init__(self,
                 path: Path = "",
                 mode: str = "s",
                 l_ext: str | list[str] = None
                 ):
        self.f = ""
        self.md = None
        self.st = None
        self.mode = mode
        self.l_ext = [".textgrid"] if l_ext is None else l_ext
        self._set_md(path)

        # Métadonnées #
        # -------------#

    def _set_md(self, path: Path | Meta = "") -> None:
        """Initialise les métadonnées."""
        if isinstance(path, Meta):
            self.f, self.md = "", path
            return
        elif (not self.md) or (path and self.f != path):
            self.f, self.md = path, Meta(path)
            self.md.load()

    def set_meta_stats(
            self, st: StList, path: Path = "",
            safe: bool = True, save: bool = True
    ):
        """Met à jour les statistiques du fichier de métadonnées."""

        fset: Callable[[str, str, str, Any], None] = self.md.ch_set \
            if safe else self.md.set
        age, wd, dur = "age", "nb_mots", "duree"
        date_enr, date_birth = "date_enregistrement", "date_naissance"
        self._set_md(path)
        for trcode, stf in st.fi.items():
            fset(trcode, "", wd, str(stf.wd))
            fset(trcode, "", dur, f"{stf.dur}:.03f")
            for spkcode in stf.spk:
                enr = datetime.datetime.strptime(
                    self.md.get(trcode, spkcode, date_enr), CELL_D)
                birth = datetime.datetime.strptime(
                    self.md.get(trcode, spkcode, date_birth), CELL_Y)
                fset(trcode, spkcode, age, str(enr.year - birth.year))
        self.md.save(self.f, True)

    def ch_typ(self, typ: str) -> Callable | None:
        """
        Vérifie si 'typ' appartient à la transcription ou au locuteur.
        Retourne la fonction correspondante.
        """
        fun = None
        if not self.md:
            return fun
        ch = self.md.ch_key(typ)
        return self.sort_tr if ch else self.sort_spk

    def sort_tr(
            self, st: StList, typ: str, path: Path = ""
    ) -> dict[str, StList]:
        """Catégorise les statistiques par 'typ' (transcription)."""
        self._set_md(path)
        res = {}
        for trcode, stf in st.fi.items():
            k: str = self.md.get(trcode, "trans", typ)
            if k not in res:
                res[k] = StList()
            res[k].fi[trcode] = stf
            fill_stlist(res[k], trcode)
        return res

    def sort_spk(
            self, st: StList, typ: str, path: Path = ""
    ) -> dict[str, StList]:
        """Catégorise les statistiques par 'typ' (locuteur)."""
        self._set_md(path)
        res = {}
        for trcode, stf in st.fi.items():
            for spkcode, tpl in stf.spk.items():
                wd, dur = tpl
                k: str = self.md.get(trcode, spkcode, typ)
                if k not in res:
                    res[k] = StList()
                if trcode not in res[k].fi:
                    res[k].fi[trcode] = StFile()
                res[k].wd += wd
                res[k].dur += dur
                nstf = res[k].fi[trcode]
                nstf.wd += wd
                nstf.dur += dur
                nstf.spk[spkcode] = tpl
                if spkcode not in res[k].spk:
                    res[k].spk[spkcode] = [0, 0.0]
                nspk = res[k].spk[spkcode]
                nspk[0] += wd
                nspk[1] += dur
        return res

        # Données #
        # ---------#

    def load_dir(self,
                 path: Path,
                 mode: str = "",
                 l_ext: str | list[str] = None
                 ) -> StList:
        """Charge les statistiques d'un dossier."""
        self.mode = mode if mode else self.mode
        self.l_ext = fix_lext(l_ext) if l_ext else self.l_ext
        self.st = all_file_stats(path, self.mode, self.l_ext)
        return self.st

    def load_corp(self,
                  corp: list[str] = None,
                  mode: str = "",
                  l_ext: str | list[str] = None
                  ) -> StList:
        """Charge les statistiques du corpus OFROM+."""
        self.mode = mode if mode else self.mode
        self.l_ext = fix_lext(l_ext) if l_ext else self.l_ext
        self.st = corp_stats(corp, self.mode, self.l_ext)
        return self.st

    def load_meta(self, path: Path) -> None:
        """Recharge les métadonnées."""
        self._set_md(path)

        # Tri #
        # -----#

    def sort(self,
             st: StList,
             typ: str,
             func: Callable = None
             ) -> dict[str, StList]:
        """Retourne les statistiques triées par 'typ'."""
        func = self.ch_typ(typ) if not func else func
        return func(st, typ)

        # Sauvegarde #
        # ------------#

    def to_excel(self,
                 path: Path,
                 d_st: dict[str, StList],
                 shn: str = "general"
                 ) -> None:
        """Sauvegarde les statistiques dans un fichier Excel."""
        wb, sh = open_excel(path, shn)
        _write_table(sh, d_st, HEADER)
        wb.save(path)

    def to_excel_typ(self,
                     path: Path,
                     st: StList,
                     typ: str
                     ) -> None:
        """
        Sauvegarde les statistiques dans un fichier Excel.
        Crée une table par sous-corpus plus une table générale.
        """
        wb, sh = open_excel(path, typ)
        func = self.ch_typ(typ)
        head = HEADER.copy()
        d_st = self.sort(st, SCOR, self.sort_tr)
        for corp, sst in d_st.items():
            head[0] = corp
            d_sst = self.sort(sst, typ, func)
            _write_table(sh, d_sst, head)
        head[0] = typ
        _write_table(sh, self.sort(st, typ, func), head)
        wb.save(path)


def get_corpus_stats(
        meta_path: Path = "",
        corp: list[str] = None,
        l_typs: list[str] = None,
        mode: str = "s",
        l_ext: str | list[str] = None,
        save_path: Path = "stats.xlsx"
) -> None:
    """
    Charge les données du corpus OFROM+ (sous-corpus 'corp', vide pour tous), 
    ajoute une page par type 
    """
    l_typs = [] if l_typs is None else l_typs
    st = Stats(meta_path)
    st.load_corp(corp, mode, l_ext)
    d_st = st.sort(st.st, SCOR)
    st.to_excel(save_path, d_st, "general")
    for typ in l_typs:
        st.to_excel_typ(save_path, st.st, typ)


if __name__ == "__main__":
    from ofrom_outils.common import META

    get_corpus_stats(
        META,
        ['OFROM_classique', 'OFROM_multigenres', 'OFROM_vocaux'],
        ['sexe', 'region', 'genre']
    )
