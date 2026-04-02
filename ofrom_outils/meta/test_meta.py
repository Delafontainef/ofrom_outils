import os
import tempfile
import unittest
from unittest.mock import patch

import openpyxl as xl
from openpyxl.workbook import Workbook

from ofrom_outils.common_types import Transcription
from ofrom_outils.meta.meta import (
    Meta
)
from ofrom_outils.meta.meta_models import (
    MetaDict, Tr, Spk
)


def set_tmp(suffix=".xlsx"):
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_name = tmp.name
    return tmp_name

    # test classes #
    # --------------#


class TestMeta(unittest.TestCase):
    """'Meta', manipulation de fichiers."""

    #### TODO: refaire tous les tests de méthodes

    def setUp(self):
        self.meta = Meta()
        self.wb = xl.Workbook()
        sh = self.wb.active
        sh.title = "sup"
        sh.append(['nom_dossier', 'code_locuteur'])
        sh.append(['t26a01', 't26_001'])

    def test_invalid_path(self):
        of = self.meta.f
        self.meta.set_path("hello world")
        self.assertEqual(self.meta.f, of)

    def test_valid_path(self):
        tmp_name = set_tmp()
        try:
            self.meta.set_path(tmp_name)
            self.assertEqual(self.meta.f, tmp_name)
        finally:
            os.remove(tmp_name)

    def test_open(self):
        tmp_name = set_tmp()
        try:
            self.wb.save(tmp_name)
            self.meta.open(tmp_name)
            self.assertEqual(self.wb.sheetnames, self.meta.wb.sheetnames)
        finally:
            os.remove(tmp_name)

    def test_close(self):
        self.meta.close()
        self.assertTrue(self.meta.wb is None)

    def test_clear(self):
        self.meta.wb = xl.Workbook()
        self.meta.d = MetaDict(["hello", "what"], [], {}, {})
        self.meta.clear()
        self.assertEqual((self.meta.wb, self.meta.d), (None, MetaDict()))

    def test_save(self):
        self.meta.wb = self.wb
        tmp_name = set_tmp()
        try:
            self.meta.save(tmp_name)
            xl.load_workbook(tmp_name)
            self.assertTrue(self.meta.wb.sheetnames, self.wb.sheetnames)
        finally:
            os.remove(tmp_name)


@patch.object(Meta, 'open')
@patch.object(Meta, 'close')
@patch.object(Meta, 'clear')
@patch.object(Meta, 'save')
class TestMetaContinued(unittest.TestCase):
    """'Meta', manipulation de données.."""

    def setUp(self):
        self.meta = Meta()
        self.md = MetaDict()
        self.md.tr['t26a01'] = Tr()
        self.md.tr['t26a01'].spk.append("t26_001")
        self.md.tr['t26a01'].d['location'] = "here"
        kspk = ('t26a01', 't26_001')
        self.md.spk[kspk] = Spk()
        self.md.spk[kspk].d['age'] = "125"
        self.md.spk[kspk].sh = ('sup', 2)
        self.md.tr_cols = ['location']
        self.md.spk_cols = ['age']

    @patch("ofrom_outils.meta.meta.load_meta")
    def test_load(self, mload, _msave, _mclear, _mclose, _mopen):
        mload.return_value = self.md
        self.meta.load("some/path")
        mload.assert_called_once()
        self.assertEqual(self.meta.d, self.md)

    def test_ch_key(self, _msave, _mclear, _mclose, _mopen):
        self.meta.d = self.md
        self.assertEqual(self.meta.ch_key('t26a01'), 'trans')
        self.assertEqual(self.meta.ch_key('t26_001'), '')
        self.assertRaises(KeyError, self.meta.ch_key, 't25')

    @patch("ofrom_outils.meta.meta.get_meta")
    def test_get(self, mget, _msave, _mclear, _mclose, _mopen):
        mget.return_value = ["t26_001"]
        res = self.meta.get("t26a01")
        mget.assert_called_once()
        self.assertEqual(res, ["t26_001"])

    @patch("ofrom_outils.meta.meta.set_meta")
    def test_set(self, mset, _msave, _mclear, _mclose, _mopen):
        mset.return_value = self.md
        self.meta.set("t26a01", "t26_001", "nom", "Petzi")
        mset.assert_called_once()
        self.assertEqual(self.meta.d, self.md)

    @patch("ofrom_outils.meta.meta.set_meta")
    def test_ch_set(self, mset, _msave, _mclear, _mclose, _mopen):
        mset.return_value = self.md
        self.meta.ch_set("t26a01", "t26_001", "nom", "Petzi")
        mset.assert_called_once()
        self.assertEqual(self.meta.d, self.md)
        mset.reset_mock()
        self.meta.ch_set("t26a01", "t26_001", "age", "18")
        assert mset.call_count == 0

    def test_tr_cols(self, _msave, _mclear, _mclose, _mopen):
        self.meta.d = self.md
        self.assertEqual(self.meta.tr_cols(), ['location'])

    def test_spk_cols(self, _msave, _mclear, _mclose, _mopen):
        self.meta.d = self.md
        self.assertEqual(self.meta.spk_cols(), ['age'])

    def test_iter_tr(self, _msave, _mclear, _mclose, _mopen):
        self.meta.d = self.md
        l_res = [tpl for tpl in self.meta.iter_tr()]
        self.assertEqual(l_res, [('t26a01', Tr(d={'location': 'here'},
                                               spk=['t26_001']))])

    def test_iter_spk(self, _msave, _mclear, _mclose, _mopen):
        self.md.tr['t26a01'].spk.append("t26_002")
        kspk = ('t26a01', 't26_002')
        self.md.spk[kspk] = Spk()
        self.md.spk[kspk].d['age'] = "8"
        self.md.tr['t26a02'] = Tr()
        self.md.tr['t26a02'].spk.append("t26_003")
        kspk = ('t26a02', 't26_003')
        self.md.spk[kspk] = Spk()
        self.meta.d = self.md
        l_res = [tpl for tpl in self.meta.iter_spk()]
        self.assertEqual(l_res, [
            ('t26a01', 't26_001', {'age': '125'}),
            ('t26a01', 't26_002', {'age': '8'}),
            ('t26a02', 't26_003', {})
        ])
        l_res = [tpl for tpl in self.meta.iter_spk('t26a02')]
        self.assertEqual(l_res, [('t26a02', 't26_003', {})])

    @patch("ofrom_outils.meta.meta.set_parent")
    def test_add_to_trans(self, msetp, _msave, _mclear, _mclose, _mopen):
        msetp.return_value = None
        self.meta.d = self.md
        trans = Transcription('t26a01')
        ptier = trans.create(0, "t26_001", 0., 10.)
        ctier = trans.create(1, "t26_001[anno]", 0., 10.)
        ctier.setParent(ptier)
        self.meta.add_to_trans(trans)
        self.assertEqual(ptier.meta('speaker'), 't26_001')
        self.assertEqual(ctier.meta('speaker'), 't26_001')
        trans.name = 't25'
        msetp.reset_mock()
        self.meta.add_to_trans(trans)
        assert msetp.call_count == 0

    @patch("ofrom_outils.meta.meta.set_pub_meta")
    @patch("ofrom_outils.meta.meta.save_as_csv")
    def test_set_pub(
            self, _mcsv, mpub, _msave, _mclear, _mclose, _mopen
    ):
        wb = xl.Workbook()
        wb.active.append(["a1", "b1"])
        wb.active.append(["1", "2"])
        tmp = set_tmp(suffix="")
        try:
            mpub.return_value = (wb, tmp)
            md = self.meta.set_pub("OFROM-test", save=True)
            xl.load_workbook(tmp + ".xlsx")
            self.assertTrue(isinstance(md, Workbook))
        finally:
            os.remove(tmp)


if __name__ == "__main__":
    unittest.main()
