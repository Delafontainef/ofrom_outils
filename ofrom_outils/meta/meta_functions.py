import csv
import datetime

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook

from ofrom_outils.common import DFLT, iter_file
from ofrom_outils.common_types import Any, Iterator, Path, Row, Worksheet
from ofrom_outils.meta.meta_models import (MetaDict, Tr, Spk)
from ofrom_outils.meta.meta_validation import VVal, VCell
from ofrom_outils.pr.private_paths import sub_corpus

type IterMeta = Iterator[tuple[str, object, dict, int, Row]]
type SpktoSh = list[tuple[str, int]]

"""Constantes globales
clé         type        description
-------------------------------------------------
TRCODE      str         Le nom de colonne pour la transcription (id)
SPKCODE     str         Le nom de colonne pour le locuteur (id)
ENQU        str         Un nom de locuteur réservé.
META_TR     list<str>   Les noms de colonnes pour la transcription.
META_SPK    list<str>   Les noms de colonnes pour le locuteur.
META_PUB    list<str>   Les noms de colonnes fournis au public
"""
TRCODE: str = "nom_dossier"
SPKCODE: str = "code_locuteur"
ENQU: str = "enqu"
META_TR: list[str] = [
    'statut', 'sous-corpus', 'responsable', 'universite',
    'lieu_enregistrement', 'region_enregistrement',
    'date_enregistrement', 'genre',
    'enqueteur', 'transcripteur', 'reviseur_1', 'reviseur_2',
    'reviseur_1_date', 'reviseur_2_date',
    'nb_mots', 'duree', 'qualite', 'doi'
]
META_SPK: list[str] = [
    'sexe', 'age', 'annee_naissance', 'pays', 'region', 'departement',
    'domicile_jeunesse', 'domicile_actuel', 'habite_depuis',
    'langage', 'metier', 'niveau_socioeducatif',
    'statut_familial', 'role', 'degre_proximite', 'nature_lien',
    'longitude', 'latitude', 'extr_deb', 'extr_fin'
]
META_PUB = [
    "sous-corpus", "nom_dossier", "code_locuteur", "sexe",
    "annee_naissance", "age", "pays", "region", "departement",
    "domicile_jeunesse", "domicile_actuel", "habite_depuis",
    "langage", "niveau_socioeducatif", "statut_familial",
    "role", "degre_proximite", "nature_lien",
    "date_enregistrement", "lieu_enregistrement",
    "region_enregistrement", "nb_mots", "duree", "qualite",
    "genre", "responsable", "universite", "doi"
]


# Métadonnées #
# -------------#
def iter_shn(wb: Workbook) -> IterMeta:
    """Itère sur le contenu du WorkBook."""
    for shn in wb.sheetnames:
        sh = wb[shn]
        d_c = {c.value: i for i, c in enumerate(sh[1])}
        for i, row in enumerate(sh.iter_rows(min_row=2)):
            yield shn, sh, d_c, i + 2, row


def _mrow(
        md: MetaDict, shn: str, i: int, row: Row, d_c: dict[str, int]
) -> MetaDict:
    """Récupère les métadonnées pour une ligne du 'metadata'."""
    trcode = VCell(row[d_c[TRCODE]], TRCODE).value
    spkcode = VCell(row[d_c[SPKCODE]], SPKCODE).value
    m_tr = [mtr for mtr in META_TR if mtr in d_c]
    m_spk = [mspk for mspk in META_SPK if mspk in d_c]
    if trcode not in md.tr:
        md.tr[trcode] = Tr()
        for mtr in m_tr:
            md.tr[trcode].d[mtr] = VCell(row[d_c[mtr]], mtr).value
        md.tr_cols = m_tr
    kspk = (trcode, spkcode)
    if spkcode not in md.tr[trcode].spk:  # new speaker
        md.tr[trcode].spk.append(spkcode)
        md.spk[kspk] = Spk()
        md.spk_cols = m_spk
    for mspk in m_spk:
        md.spk[kspk].d[mspk] = VCell(row[d_c[mspk]], mspk).value
    md.spk[kspk].sh = (shn, i)
    return md


def load_meta(wb: Workbook) -> MetaDict:
    """Charge le dictionnaire de métadonnées 'MetaDict'."""
    md = MetaDict()
    for shn, sh, d_c, i, row in iter_shn(wb):
        _mrow(md, shn, i, row, d_c)
    return md


def get_meta(
        md: MetaDict,
        trcode: str, spkcode: str = 'trans', k: str = ""
) -> str | list | dict[str, str]:
    """
    Récupère une ou des métadonnées.
    - md        MetaDict    Les métadonnées en mémoire
    - trcode    str         Le nom de la transcription
    - spkcode   str         Le nom du locuteur
    - k         str         Le nom de la métadonnée
    Par défaut renvoie toutes les données de la transcription.
    > spkcode = '' pour la liste des locuteurs de la transcription.
    > spkcode valide pour toutes les métadonnées du locuteur.
    > k valide pour une métadonnée précise.
    """
    if ENQU in spkcode:  # cas de l'enquêteur
        return DFLT if k else {ck: DFLT for ck in md.spk_cols}
    if trcode not in md.tr:
        raise KeyError(f"{trcode} pas dans 'MetaDict'.")
    if not spkcode:  # liste des locuteurs
        return md.tr[trcode].spk.copy()
    kspk = (trcode, spkcode)
    dat = md.tr[trcode].d if spkcode not in md.tr[trcode].spk else \
        md.spk[kspk].d
    if k and k not in dat:  # nom de métadonnée invalide
        raise KeyError(f"{kspk}: '{k}' pas dans 'MetaDict'.")
    return dat[k] if k else dat.copy()  # donnée unique ou paquet


def _set_meta_md(
        md: MetaDict, trcode: str, spkcode: str, k: str, v: Any
) -> SpktoSh:
    """
    Modifie dictionnaire de métadonnées.
    Retourne les lignes à éditer dans le WorkBook.
    """
    if trcode not in md.tr:  # rien à changer
        return []
    elif spkcode in md.tr[trcode].spk:  # métadonnée de locuteur
        l_spk = [md.spk[(trcode, spkcode)].sh]
        md.spk[(trcode, spkcode)].d[k] = VVal(v).value
    else:  # métadonnée de transcription
        l_spk = [md.spk[(trcode, spk)].sh
                 for spk in md.tr[trcode].spk]
        md.tr[trcode].d[k] = VVal(v).value
    return l_spk


def _set_meta_wb(
        wb: Workbook, l_spk: SpktoSh,
        trcode: str, spkcode: str, k: str, v: Any
) -> None:
    """Modifie le WorkBook (sans sauvegarder)."""
    sh, oshn, d_c = None, "", {}
    for shn, i in l_spk:
        if shn not in wb.sheetnames:  # perdu la trace des métadonnées...
            raise KeyError(f"{(trcode, spkcode)}: '{shn}'"
                           f" pas dans le fichier.")
        elif i < 2:
            raise IndexError(f"{(trcode, spkcode)}: '{shn}-{i}' "
                             f"ligne incorrecte.")
        elif oshn != shn:  # nouvelle feuille
            sh, oshn = wb[shn], shn
            d_c = {c.value: j + 1 for j, c in enumerate(sh[1])}
        if k not in d_c:  # pas dans l'en-tête...
            raise KeyError(f"{(trcode, spkcode)}: '{k}'"
                           f" pas dans le fichier.")
        VCell(sh.cell(int(i), d_c[k])).value = v


def set_meta(
        wb: Workbook, md: MetaDict,
        trcode: str, spkcode: str, k: str, v: Any
) -> MetaDict:
    """Modifie dictionnaire et WorkBook (sans sauvegarder)."""
    l_spk = _set_meta_md(md, trcode, spkcode, k, v)  # modifie 'MetaEdit'
    _set_meta_wb(wb, l_spk, trcode, spkcode, k, v)  # modifie WorkBook
    return md


def save_as_csv(sh: Worksheet, path: Path) -> None:
    """Sauvegarde le metadata comme '.csv'."""
    with open(path, "w", newline="", encoding="utf-8") as wf:
        w = csv.writer(wf)
        for row in sh.iter_rows(values_only=True):
            w.writerow(row)

    # Métadonnées publiques #
    # -----------------------#


def get_pub_files(c_path: Path) -> dict[str, Path]:
    """Génère la liste des fichiers publics comme dictionnaire."""
    d_files = {}
    for fi, ext, file, path in iter_file(c_path):
        d_files[fi] = path
    return d_files


def set_pub_meta(wb: Workbook, corp: str) -> tuple[Workbook, Path]:
    """
    Génère un metadata public à partir d'une feuille et 
    d'un dictionnaire de fichiers.
    Note : seules les entrées avec un fichier correspondant sont partagées.
    """
    # mise en place
    corp = corp.replace("-", "_")  # par prudence...
    sh = wb[corp.replace("_", "-")]
    c_path = sub_corpus(corp)
    d_files = get_pub_files(c_path)
    d_c = {c.value: i for i, c in enumerate(sh[1])}
    # écriture
    nwb = xl.Workbook()
    nsh = nwb.active
    nsh.title = sh.title
    nsh.append(META_PUB)
    for row in sh.iter_rows(min_row=2):
        fi = row[d_c[TRCODE]].value
        if fi not in d_files:  # uniquement métadonnées publiques
            continue
        nrow = [row[d_c[col]].value if col in d_c else ""
                for col in META_PUB]  # tous les champs devraient y être
        # formatage des dates de dernière minute...
        for i, v in enumerate(nrow):
            try:
                parsed = datetime.datetime.strptime(v, "%Y-%m-%d")
                nrow[i] = parsed.strftime("%d-%m-%Y")
            except ValueError:
                continue
        nsh.append(nrow)
    return nwb, c_path
