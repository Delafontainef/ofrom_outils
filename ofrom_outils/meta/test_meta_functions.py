from ofrom_outils.meta.meta_models import (MetaDict, Tr, Spk)
from ofrom_outils.meta.meta_functions import (
    iter_shn, _mrow, load_meta, get_meta, set_meta, 
    _set_meta_md, _set_meta_wb, save_as_csv,
    get_pub_files, set_pub_meta
)
from ofrom_outils.meta.meta_validation import VCell
import unittest
from unittest.mock import patch, Mock
import os, tempfile, copy
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
import openpyxl as xl
import warnings
warnings.simplefilter("ignore")

def set_tmp(suffix=".xlsx"):
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_name = tmp.name
    return tmp_name
def set_wb():
    wb = xl.Workbook()
    wb.active.title = "sh1"
    wb.create_sheet("sh2")
    for shn in wb.sheetnames:
        wb[shn].append(['nom_dossier', 'code_locuteur', 
                        'nom', 'age', 'universite'])
    wb['sh1'].append(['equ01', 'equ_001', 'TS', '6', 'CH'])
    wb['sh2'].append(['equ02', 'equ_002', 'RD', '5', 'WB'])
    wb['sh2'].append(['equ02', 'equ_003', 'AJ', '6', 'WB'])
    return wb
def set_md():
    md = MetaDict()
    tr = md.tr['equ01'] = Tr()
    tr.d, tr.spk = {'universite': 'CH'}, ['equ_001']
    kspk = ('equ01', 'equ_001')
    spk = md.spk[kspk] = Spk()
    spk.d, spk.sh = {'age': '6'}, ('sh1', 2)
    tr = md.tr['equ02'] = Tr()
    tr.d, tr.spk = {'universite': 'WB'}, ['equ_002', 'equ_003']
    kspk = ('equ02', 'equ_002')
    spk = md.spk[kspk] = Spk()
    spk.d, spk.sh = {'age': '5'}, ('sh2', 2)
    kspk = ('equ02', 'equ_003')
    spk = md.spk[kspk] = Spk()
    spk.d, spk.sh = {'age': '6'}, ('sh2', 3)
    return md
WB = set_wb()   # "mock" Workbook
MD = set_md()   # "mock" MetaDict

    # test classes #
    #--------------#
class TestIterShn(unittest.TestCase):
    
    def test_iter(self):
        l_res = [res for res in iter_shn(WB)]
        self.assertEqual(len(l_res), 3)
        self.assertEqual(l_res[1][0], l_res[2][0])
        self.assertEqual(l_res[2][2]['nom'], 2)
        self.assertTrue(isinstance(l_res[0][4][0], Cell))
    def test_iter_fail(self):
        with self.assertRaises(AttributeError):
            list(iter_shn("hello"))
        
class TestMRow(unittest.TestCase):
    
    def setUp(self):
        self.row = WB['sh1'][2]
        self.d_c = {k.value: i for i, k in enumerate(WB['sh1'][1])}
    
    def test_valid(self):
        md = _mrow(MetaDict(), 'sh1', 2, self.row, self.d_c)
        self.assertEqual(md.tr['equ01'], MD.tr['equ01'])
        kspk = ('equ01', 'equ_001')
        self.assertEqual(md.spk[kspk], MD.spk[kspk])  
    def test_invalid(self):
        self.assertRaises(KeyError, _mrow, MetaDict(), 'sh1', 1, self.row, {})
        self.assertRaises(AttributeError, _mrow, None, 
                          '', 1, self.row, self.d_c)
        self.assertRaises(IndexError, _mrow, MetaDict(), 'sh1', 1, "", self.d_c)

@patch("ofrom_outils.meta.meta_functions._mrow")
@patch("ofrom_outils.meta.meta_functions.iter_shn")
class TestLoadMeta(unittest.TestCase):
    
    def setUp(self):
        self.md = MetaDict()
    
    def test_load(self, mock_iter, mock_mrow):
        mock_iter.return_value = [(1, 2, 3, 4, 5), (6, 7, 8, 9, 10)]
        md = load_meta(WB)
        mock_iter.assert_called_once_with(WB)
        self.assertEqual(mock_mrow.call_count, 2)
        mock_mrow.assert_any_call(md, 1, 4, 5, 3)

class TestGetMeta(unittest.TestCase):
    
    def test_no_trcode(self):
        self.assertRaises(KeyError, get_meta, MD, "")
    def test_dflt_spkcode(self):
        self.assertEqual(get_meta(MD, "equ01"), MD.tr['equ01'].d)
    def test_no_spkcode(self):
        self.assertEqual(get_meta(MD, "equ01", ""), MD.tr['equ01'].spk)
    def test_speaker(self):
        self.assertEqual(get_meta(MD, "equ01", "equ_001"), 
                         MD.spk[('equ01', 'equ_001')].d)
    def test_tr_arg(self):
        self.assertEqual(get_meta(MD, "equ01", "ah ah", "universite"),
                         MD.tr['equ01'].d['universite'])
    def test_spk_arg(self):
        self.assertEqual(get_meta(MD, "equ01", "equ_001", "age"),
                         MD.spk[('equ01', 'equ_001')].d['age'])
    def test_arg_fail(self):
        self.assertRaises(KeyError, get_meta, MD, "equ01", "prrrt", "age")

class TestSetMetaMD(unittest.TestCase):
    
    def setUp(self):
        self.md = copy.deepcopy(MD) # ne pas changer au global
        
    def test_invalid_trcode(self):
        self.assertEqual(_set_meta_md(self.md, "lol", "spk", "k", "v"), 
                         self.md)
    def test_spk_md(self):
        l_spk = _set_meta_md(self.md, "equ02", "equ_003", "ch", "63")
        kspk = ("equ02", "equ_003")
        self.assertEqual(self.md.spk[kspk].d['ch'], "63")
        with self.assertRaises(KeyError):
            v = self.md.spk[('equ02', 'equ_002')].d['ch']
        self.assertEqual(l_spk, [('sh2', 3)])
    def test_tr_md(self):
        l_spk = _set_meta_md(self.md, "equ02", "lol", "check", "True")
        self.assertEqual(self.md.tr['equ02'].d['check'], "True")
        self.assertEqual(l_spk, [('sh2', 2), ('sh2', 3)])
    
class TestSetMetaWB(unittest.TestCase):
        
    def test_invalid_sheet(self):
        self.assertRaises(KeyError, _set_meta_wb,
                          WB, [('sh0', 2)], "", "", "", "")
    def test_invalid_key(self):
        self.assertRaises(KeyError, _set_meta_wb,
                          WB, [('sh1', 2)], "", "", "opal", "yes")
    @patch("ofrom_outils.meta.meta_functions.VCell")
    def test_valid(self, Mock_vc):
        mock = Mock_vc.return_value
        _set_meta_wb(WB, [('sh1', 2)], "", "", "age", 7)
        Mock_vc.assert_called_once()

@patch("ofrom_outils.meta.meta_functions._set_meta_md")
@patch("ofrom_outils.meta.meta_functions._set_meta_wb")
class TestSetMeta(unittest.TestCase):
    
    def test_valid(self, mock_wb, mock_md):
        mock_md.return_value = [('sh1', 2)]
        res = set_meta(WB, MD, "equ01", "equ_001", "age", 7)
        mock_md.assert_called_once()
        mock_wb.assert_called_once()
        self.assertEqual(MD, res)

class TestSaveAsCSV(unittest.TestCase):
    
    def test_save_csv(self):
        tmp = set_tmp(".csv")
        try:
            save_as_csv(WB['sh2'], tmp)
            with open(tmp, "r", encoding="utf8") as rf:
                line = rf.readline()
                self.assertTrue(line.startswith("nom_dossier,code_locuteur"))
        finally:
            os.remove(tmp)

@patch("ofrom_outils.meta.meta_functions.iter_file")
class TestGetPubFiles(unittest.TestCase):
    
    def test_get_files(self, mock_iter):
        mock_iter.return_value = [('fi1', 'ext', 'file', 'p1'),
                                  ('fi2', 'ext', 'file', 'p2')]
        d_fi = get_pub_files("test")
        mock_iter.assert_called_once_with("test")
        self.assertEqual(d_fi, {'fi1': 'p1', 'fi2': 'p2'})

@patch("ofrom_outils.meta.meta_functions.get_pub_files")
class TestSetPubMeta(unittest.TestCase):
    
    def test_set_pub(self, mock_getf):
        mock_getf.return_value = {'equ01': '', 'equ02': ''}
        nwb, c = set_pub_meta(WB, "sh2", "pubf")
        d_c = {k.value: i for i, k in enumerate(nwb['sh2'][1])}
        mock_getf.assert_called_once()
        self.assertEqual(nwb['sh2'][2][d_c['age']].value, "5")
        
    def test_set_pub_date(self, mock_getf):
        mock_getf.return_value = {'equ01': '', 'equ02': ''}
        WB['sh2'].insert_cols(1)
        for i, el in enumerate(["date_enregistrement", 
                                "2011-10-28", "2012-05-14"]):
                WB['sh2'][i+1][0].value = el
        nwb, c = set_pub_meta(WB, "sh2", "pubf")
        d_c = {k.value: i for i, k in enumerate(nwb['sh2'][1])}
        self.assertEqual(nwb['sh2'][2][d_c['date_enregistrement']].value, 
                         "28-10-2011")

if __name__ == "__main__":
    unittest.main()